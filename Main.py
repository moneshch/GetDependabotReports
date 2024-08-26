import os
import logging
from datetime import datetime
import requests
import pandas as pd
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import GetDf
import openpyxl

# Suppress InsecureRequestWarning caused by disabling SSL verification
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Set up logging
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(log_dir, f"log_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")
logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# File paths

output_file_name = 'Output_' + datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + '.xlsx'
script_dir = os.path.dirname(__file__)
output_file_path = os.path.join(script_dir, "Output", output_file_name)
logging.info(f"Input read")


# Create an empty DataFrame and write it to the output Excel file
empty_df = pd.DataFrame()
empty_df.to_excel(output_file_path, index=False)
logging.info(f"Empty file created")
# Iterate through each row in the DataFrame

print("Enter Organization name")

org_name = input()

print("Enter Fine-grained personal access tokens")
token = input()

headers = {'Authorization': f'token {token}'}

urlGetAlert = f'https://api.github.com/orgs/{org_name}/dependabot/alerts'
logging.info(f" url is {urlGetAlert}")
resAlerts = requests.get(urlGetAlert, headers=headers, verify=False)
logging.info(f" response is {resAlerts}")
dffAlerts = []

if (resAlerts.status_code == 200 and (resAlerts.json())):
        dataAlerts = []
        page = 1

        while True:
            urlGetAlert2 = f'https://api.github.com/orgs/{org_name}/dependabot/alerts?per_page=100&page={page}'
            logging.info(f" url is {urlGetAlert}")
            resAlerts2 = requests.get(urlGetAlert2, headers=headers, verify=False, timeout=5)
            logging.info(f" response is {resAlerts2}")

            if resAlerts2.json():
                dataAlerts.append(resAlerts2.json())
                page += 1
            else:
                break

        for i in dataAlerts:
            dffAlerts.append(GetDf.getDf(i))

        # Concatenate all DataFrames in dffAlerts into a single DataFrame
        dffAlerts = pd.concat(dffAlerts, ignore_index=True)

        # Write dffAlerts to Excel
        logging.info(f" writiing Alerts of {org_name} to {output_file_path}  ")
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            dffAlerts.to_excel(writer, sheet_name=org_name, index=False)
        logging.info(f"Alerts writen of {org_name} to {output_file_path}  ")

        # Create pivot table for dffAlerts
        pivot_df = dffAlerts.pivot_table(index='repository_name', columns='security_vulnerability_severity', aggfunc='size', fill_value=0)
        for severity in ['critical', 'high', 'medium', 'low', 'info']:
            if severity not in pivot_df.columns:
                pivot_df[severity] = 0

        # Calculate the total count of severities for each repository_name
        pivot_df['count'] = pivot_df.sum(axis=1)
        pivot_df.reset_index(inplace=True)
        logging.info(f" count of alerts for each repository_name {pivot_df}  ")

        # Fetch repository data
        #urlRepos = f'https://api.github.com/orgs/{org_name}/repos'
        #resRepos = requests.get(urlRepos, headers=headers, verify=False)
        #dataRepos = resRepos.json()

        dataRepos=[]
        pageRepos=1
        while True:
            urlRepos = f'https://api.github.com/orgs/{org_name}/repos?per_page=100&page={pageRepos}'
            logging.info(f" url is {urlRepos}. Page {pageRepos}. ")
            resRepos = requests.get(urlRepos, headers=headers, verify=False, timeout=5)
            logging.info(f" response is {resRepos}")

            if resRepos.json():
                dataRepos.append(resRepos.json())
                pageRepos += 1
            else:
                break

        
        logging.info(f"All repos are ")
        logging.info(dataRepos)

        # Prepare dictionary to hold repository_name and HTTP Status Code
        myDict = {}
        
        for j in dataRepos:
            for i in j:
                
                #strRepo = str(i.get('name'))
                strRepo = str(i.get('name'))
                print(strRepo)
                urlGetVul = f'https://api.github.com/repos/{org_name}/{strRepo}/vulnerability-alerts'
                resVul = requests.get(urlGetVul, headers=headers, verify=False)
                myDict[strRepo] = resVul.status_code
                logging.info(f" repo {strRepo} status code for vulnerability alert is {resVul.status_code}")

        # Create DataFrame from myDict
        dfRepoCode = pd.DataFrame(list(myDict.items()), columns=['repository_name', 'HTTP Status Code'])
        logging.info(f" check if vulnerability-alerts enabled {dfRepoCode}  ")

        # Define columns for the new DataFrame to include repository_name and six other columns
        columns = ['repository_name', 'critical', 'high', 'low', 'medium', 'info', 'count']
        dfRepos = pd.DataFrame(columns=columns)

        for index, row in dfRepoCode.iterrows():
            repo_name = row['repository_name']
            http_code = row['HTTP Status Code']
            logging.info(f"check reponame {repo_name} in {pivot_df['repository_name'].values})")
            if repo_name not in pivot_df['repository_name'].values:
                
                logging.info(f" reponame {repo_name} not found")

                if http_code == 404:
                    values = ['Not enabled', 'Not enabled', 'Not enabled', 'Not enabled', 'Not enabled', '0']
                elif http_code == 204:
                    values = [0, 0, 0, 0, 0, 0]
                else:
                    values = ['Unauthorized', 'Unauthorized', 'Unauthorized', 'Unauthorized', 'Unauthorized', '0']

                # Create a new DataFrame with the new row
                new_row = pd.DataFrame([[repo_name] + values], columns=pivot_df.columns)
                logging.info(f" new summary row is {new_row}  ")

                # Append the new row to dfRepos
                pivot_df = pd.concat([pivot_df, new_row], ignore_index=True)

        # Write dfRepos to Excel
        logging.info(f" final summary is {pivot_df}  ")
        logging.info(f" Writting summary  ")
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            pivot_df.to_excel(writer, sheet_name=org_name + '_Summary', index=False)
        logging.info(f"  summary  written ")
        workbook_path = output_file_path
        wb = openpyxl.load_workbook(workbook_path)

        # Delete sheet1
        if 'Sheet1' in wb.sheetnames:
            del wb['Sheet1']

        # Move summary to the first position
        summarySheet=org_name + '_Summary'
        if summarySheet in wb.sheetnames:
            sheet3 = wb[summarySheet]
            wb._sheets.remove(sheet3)
            wb._sheets.insert(0, sheet3)

        # Save the workbook
        wb.save(workbook_path)


        # Logging
        logging.info(f"Processed data for organization {org_name}")
        print(f"Processed data for organization {org_name}")

elif (resAlerts.status_code == 404):
        print(f" Organization - {org_name} not found")
        logging.info(f" Organization -  {org_name} not found")
        
elif (resAlerts.status_code == 401):
        print(f" Unauthorized ")
        logging.info(f" Unauthorized ")
else:
        print('No alerts')
        # Fetch repository data
        #urlRepos = f'https://api.github.com/orgs/{org_name}/repos'
        #resRepos = requests.get(urlRepos, headers=headers, verify=False)
        #dataRepos = resRepos.json()
        dataRepos=[]
        pageRepos=1
        while True:
            urlRepos = f'https://api.github.com/orgs/{org_name}/repos?per_page=100&page={pageRepos}'
            logging.info(f" url is {urlRepos}. Page {pageRepos}. ")
            resRepos = requests.get(urlRepos, headers=headers, verify=False, timeout=5)
            logging.info(f" response is {resRepos}")
            #print(resRepos.json())

            if resRepos.json():
                dataRepos.append(resRepos.json())
                pageRepos += 1
            else:
                break
        # Prepare dictionary to hold repository_name and HTTP Status Code
        myDict = {}
        for j in dataRepos:
            for i in j:
                
                #strRepo = str(i.get('name'))
                strRepo = str(i.get('name'))
                
                print(strRepo)
                urlGetVul = f'https://api.github.com/repos/{org_name}/{strRepo}/vulnerability-alerts'
                resVul = requests.get(urlGetVul, headers=headers, verify=False)
                myDict[strRepo] = resVul.status_code
                logging.info(f" repo {strRepo} status code for vulnerability alert is {resVul.status_code}")
            

        # Create DataFrame from myDict
        dfRepoCode = pd.DataFrame(list(myDict.items()), columns=['repository_name', 'HTTP Status Code'])

        # Define columns for the new DataFrame to include repository_name and six other columns
        columns = ['repository_name', 'critical', 'high', 'low', 'medium', 'info', 'count']
        dfRepos = pd.DataFrame(columns=columns)

        for index, row in dfRepoCode.iterrows():
            repo_name = row['repository_name']
            http_code = row['HTTP Status Code']

            if http_code == 404:
                values = ['Not enabled', 'Not enabled', 'Not enabled', 'Not enabled', 'Not enabled', '0']
            elif http_code == 204:
                values = [0, 0, 0, 0, 0, 0]
            else:
                values = ['Unauthorized', 'Unauthorized', 'Unauthorized', 'Unauthorized', 'Unauthorized', '0']

            # Create a new DataFrame with the new row
            new_row = pd.DataFrame([[repo_name] + values], columns=dfRepos.columns)

            # Append the new row to dfRepos
            dfRepos = pd.concat([dfRepos, new_row], ignore_index=True)

        # Write dfRepos to Excel
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            dfRepos.to_excel(writer, sheet_name=org_name + '_Summary', index=False)
        # Load the workbook
        workbook_path = output_file_path
        wb = openpyxl.load_workbook(workbook_path)

        # Delete sheet1
        if 'Sheet1' in wb.sheetnames:
            del wb['Sheet1']

        # Move summary to the first position
        summarySheet=org_name + '_Summary'
        if summarySheet in wb.sheetnames:
            sheet3 = wb[summarySheet]
            wb._sheets.remove(sheet3)
            wb._sheets.insert(0, sheet3)

        # Save the workbook
        wb.save(workbook_path)

        # Logging
        logging.info(f"Processed data for organization {org_name}")
        print(f"Processed data for organization {org_name}")




print("Enter any key to abort")
input()
